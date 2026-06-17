import streamlit as st
import pandas as pd
import numpy as np
import io
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import unicodedata

# ===================== 全局工具函数 =====================
def display_width(text):
    """计算单元格文本宽度，适配中英文"""
    return sum(2 if unicodedata.east_asian_width(c) in ('F','W') else 1 for c in str(text or ''))

def auto_fit_columns(ws, min_w=8, max_w=50, padding=3):
    """自动自适应列宽，无MergedCell判断避免报错"""
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        w = max((display_width(c.value) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[letter].width = max(min_w, min(w * 1.1 + padding, max_w))

def extract_asin_from_targeting(text):
    """Python端精准提取双引号内的ASIN值，适配所有格式"""
    if pd.isna(text) or str(text).strip() == "":
        return ""
    match = re.search(r'asin\s*=\s*[“"](.*?)[”"]', str(text), re.IGNORECASE)
    if match:
        return match.group(1).strip()
    match_fallback = re.search(r'[“"](.*?)[”"]', str(text))
    if match_fallback:
        return match_fallback.group(1).strip()
    return ""

# Excel全局格式定义
HEADER_FILL = PatternFill('solid', fgColor='F2F2F2')
HEADER_FONT = Font(bold=True, size=11)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(top=Side('thin'), bottom=Side('thin'), left=Side('thin'), right=Side('thin'))
NUMBER_FMT = '0.00'
PCT_FMT = '0.00%'

# ===================== 初始化默认配置【条件调整+新增精准广告高ACOS工作表】 =====================
def get_default_sheet_config():
    # 通用输出字段（前两个Sheet共用）
    common_output_cols = [
        'ASIN (Informational only)',
        'SKU',
        'Campaign Name (Informational only)',
        'Bidding Strategy',
        'Placement',
        'Percentage',
        'Impressions',
        'Clicks',
        'Click-through Rate',
        'Spend',
        'Sales',
        'Orders',
        'Units',
        'Conversion Rate',
        'ACOS',
        'CPC',
        'ROAS'
    ]
    # 定位广告类Sheet通用输出字段（永久移除Placement、Percentage列）
    targeting_output_cols = [
        'ASIN (Informational only)',
        'SKU',
        'Campaign Name (Informational only)',
        'Campaign State (Informational only)',
        'Ad Group Name (Informational only)',
        'State',
        'Bidding Strategy',
        'Product Targeting Expression',
        'Bid',
        'Impressions',
        'Clicks',
        'Click-through Rate',
        'Spend',
        'Sales',
        'Orders',
        'Units',
        'Conversion Rate',
        'ACOS',
        'CPC',
        'ROAS'
    ]
    # 精准广告类Sheet通用输出字段（严格按指定顺序）
    keyword_output_cols = [
        'ASIN (Informational only)',
        'SKU',
        'Campaign Name (Informational only)',
        'Ad Group Name (Informational only)',
        'State',
        'Bidding Strategy',
        'Keyword Text',
        'Campaign State (Informational only)',
        'Bid',
        'Impressions',
        'Clicks',
        'Click-through Rate',
        'Spend',
        'Sales',
        'Orders',
        'Units',
        'Conversion Rate',
        'ACOS',
        'CPC',
        'ROAS'
    ]
    return [
        # Sheet1：广告位无效点击
        {
            "sheet_name": "广告位无效点击",
            "filter_rules": [
                {"col": "Entity", "op": "等于", "val": "Bidding Adjustment"},
                {"col": "ACOS", "op": "等于", "val": "0"},
                {"col": "Clicks", "op": "大于", "val": "10"}
            ],
            "output_cols": common_output_cols,
            "match_campaign_asin_sku": True,
            "sort_by": "Spend",
            "sort_ascending": False,
            "is_targeting_sheet": False,
            "only_match_bidding": False
        },
        # Sheet2：广告位ACOS高
        {
            "sheet_name": "广告位ACOS高",
            "filter_rules": [
                {"col": "Entity", "op": "等于", "val": "Bidding Adjustment"},
                {"col": "ACOS", "op": "大于", "val": "0.4"}
            ],
            "output_cols": common_output_cols,
            "match_campaign_asin_sku": True,
            "sort_by": "Spend",
            "sort_ascending": False,
            "is_targeting_sheet": False,
            "only_match_bidding": False
        },
        # Sheet3：定位广告无效点击
        {
            "sheet_name": "定位广告无效点击",
            "filter_rules": [
                {"col": "Entity", "op": "等于", "val": "Product Targeting"},
                {"col": "Product Targeting Expression", "op": "包含", "val": "asin"},
                {"col": "ACOS", "op": "等于", "val": "0"},
                {"col": "Clicks", "op": "大于", "val": "10"}
            ],
            "output_cols": targeting_output_cols,
            "match_campaign_asin_sku": True,
            "sort_by": "Spend",
            "sort_ascending": False,
            "is_targeting_sheet": True,
            "only_match_bidding": True
        },
        # Sheet4：定位广告高ACOS
        {
            "sheet_name": "定位广告高ACOS",
            "filter_rules": [
                {"col": "Entity", "op": "等于", "val": "Product Targeting"},
                {"col": "Product Targeting Expression", "op": "包含", "val": "asin"},
                {"col": "ACOS", "op": "大于", "val": "0.4"}
            ],
            "output_cols": targeting_output_cols,
            "match_campaign_asin_sku": True,
            "sort_by": "Spend",
            "sort_ascending": False,
            "is_targeting_sheet": True,
            "only_match_bidding": True
        },
        # Sheet5：定位广告0点击低曝光
        {
            "sheet_name": "定位广告0点击低曝光",
            "filter_rules": [
                {"col": "Entity", "op": "等于", "val": "Product Targeting"},
                {"col": "Product Targeting Expression", "op": "包含", "val": "asin"},
                {"col": "Impressions", "op": "大于等于", "val": "0"},
                {"col": "Impressions", "op": "小于等于", "val": "100"},
                {"col": "Clicks", "op": "等于", "val": "0"}
            ],
            "output_cols": targeting_output_cols,
            "match_campaign_asin_sku": True,
            "sort_by": "Impressions",
            "sort_ascending": False,
            "is_targeting_sheet": True,
            "only_match_bidding": True
        },
        # Sheet6：精准广告无效点击【条件修改：Clicks≥5】
        {
            "sheet_name": "精准广告无效点击",
            "filter_rules": [
                {"col": "Entity", "op": "等于", "val": "Keyword"},
                {"col": "Match Type", "op": "等于", "val": "Exact"},
                {"col": "ACOS", "op": "等于", "val": "0"},
                {"col": "Clicks", "op": "大于等于", "val": "5"}
            ],
            "output_cols": keyword_output_cols,
            "match_campaign_asin_sku": True,
            "sort_by": "Spend",
            "sort_ascending": False,
            "is_targeting_sheet": False,
            "only_match_bidding": False
        },
        # Sheet7：精准广告高ACOS【新增工作表】
        {
            "sheet_name": "精准广告高ACOS",
            "filter_rules": [
                {"col": "Entity", "op": "等于", "val": "Keyword"},
                {"col": "Match Type", "op": "等于", "val": "Exact"},
                {"col": "ACOS", "op": "大于", "val": "0.4"}
            ],
            "output_cols": keyword_output_cols,
            "match_campaign_asin_sku": True,
            "sort_by": "Spend",
            "sort_ascending": False,
            "is_targeting_sheet": False,
            "only_match_bidding": False
        }
    ]

# 会话初始化
st.set_page_config(page_title="亚马逊SP广告自定义报表工具 v4.1", layout="wide")
if 'sp_df' not in st.session_state:
    st.session_state.sp_df = None
if 'sheet_config_list' not in st.session_state:
    st.session_state.sheet_config_list = get_default_sheet_config()

# ===================== 页面标题与安全提示 =====================
st.title("📊 亚马逊SP广告自定义报表生成工具")
st.caption("数据源：Sponsored Products Campaigns 广告报表 | 固定内置匹配逻辑，无需手动开关")
st.info("🛡️ **数据安全承诺**：所有文件仅本地内存处理，不上传服务器，关闭页面数据立即销毁")
st.markdown("---")

# ===================== 左侧侧边栏配置面板【Sheet默认关闭展开】 =====================
with st.sidebar:
    st.header("⚙️ Sheet自定义配置面板")
    st.divider()

    # 新增Sheet按钮
    add_sheet_btn = st.button("➕ 新增一个分析Sheet", use_container_width=True)
    if add_sheet_btn:
        st.session_state.sheet_config_list.append({
            "sheet_name": "新建Sheet",
            "filter_rules": [],
            "output_cols": [],
            "match_campaign_asin_sku": True,
            "sort_by": "Spend",
            "sort_ascending": False,
            "is_targeting_sheet": False,
            "only_match_bidding": False
        })

    # 循环渲染每个Sheet配置【默认关闭展开】
    delete_idx_list = []
    for idx, sheet_cfg in enumerate(st.session_state.sheet_config_list):
        with st.expander(f"Sheet {idx+1}：{sheet_cfg['sheet_name']}", expanded=False):
            # 1. 修改Sheet名称
            sheet_cfg["sheet_name"] = st.text_input("Sheet工作表名称", value=sheet_cfg["sheet_name"], key=f"sheet_name_{idx}")

            # 固定逻辑提示
            st.caption("固定逻辑：自动按CampaignID匹配多ASIN/SKU分行输出；定位广告Sheet自动生成亚马逊超链接、仅匹配Bidding Strategy字段")
            st.divider()

            # 2. 排序配置
            st.subheader("排序配置")
            sort_col_in, sort_order_in = st.columns([3, 2])
            sheet_cfg["sort_by"] = sort_col_in.text_input("排序字段", value=sheet_cfg["sort_by"], key=f"sort_by_{idx}")
            sort_order = sort_order_in.selectbox("排序顺序", ["降序", "升序"], index=0 if not sheet_cfg["sort_ascending"] else 1, key=f"sort_order_{idx}")
            sheet_cfg["sort_ascending"] = sort_order == "升序"

            # 3. 筛选条件配置
            st.subheader("筛选条件配置")
            add_filter_btn = st.button(f"添加筛选条件 #{idx}", key=f"add_filter_{idx}")
            if add_filter_btn:
                sheet_cfg["filter_rules"].append({"col": "", "op": "等于", "val": ""})
            del_filter_idx = []
            for f_idx, rule in enumerate(sheet_cfg["filter_rules"]):
                col_in, op_in, val_in = st.columns([3, 2, 3])
                rule["col"] = col_in.text_input("字段", value=rule["col"], key=f"f_col_{idx}_{f_idx}")
                rule["op"] = op_in.selectbox("运算符", ["等于", "大于", "小于", "大于等于", "小于等于", "包含"], index=["等于", "大于", "小于", "大于等于", "小于等于", "包含"].index(rule["op"]), key=f"f_op_{idx}_{f_idx}")
                rule["val"] = val_in.text_input("值", value=rule["val"], key=f"f_val_{idx}_{f_idx}")
                if st.button(f"删除本条条件", key=f"del_f_{idx}_{f_idx}"):
                    del_filter_idx.append(f_idx)
            for f_idx in sorted(del_filter_idx, reverse=True):
                del sheet_cfg["filter_rules"][f_idx]

            # 4. 输出表头配置
            st.subheader("输出表头（逗号分隔字段）")
            cols_text = ", ".join(sheet_cfg["output_cols"])
            input_cols_text = st.text_area("填写输出字段，英文逗号分隔", value=cols_text, height=120, key=f"output_cols_{idx}")
            clean_text = input_cols_text.replace("\n", "").replace("\t", "")
            sheet_cfg["output_cols"] = [c.strip() for c in clean_text.split(",") if c.strip()]

            # 5. 删除当前Sheet按钮
            if st.button(f"🗑️ 删除【{sheet_cfg['sheet_name']}】", type="secondary", key=f"del_sheet_{idx}"):
                delete_idx_list.append(idx)
    for del_idx in sorted(delete_idx_list, reverse=True):
        del st.session_state.sheet_config_list[del_idx]

    st.divider()
    st.info("操作说明：\n1. 可自定义筛选、排序、输出字段\n2. 内置固定逻辑无需手动开关\n3. 定位广告报表不含Placement/Percentage列")

# ===================== 文件上传模块 =====================
st.header("📁 上传亚马逊SP广告 xlsx 报表")
upload_file = st.file_uploader("仅支持包含 Sponsored Products Campaigns 工作表的xlsx文件", type=["xlsx"], key="sp_upload")

if upload_file:
    try:
        all_sheet_dict = pd.read_excel(upload_file, sheet_name=None)
        sheet_name_list = list(all_sheet_dict.keys())
        if "Sponsored Products Campaigns" not in sheet_name_list:
            st.error("❌ 文件缺少工作表：Sponsored Products Campaigns，请核对报表！")
            st.stop()
        df_sp = all_sheet_dict["Sponsored Products Campaigns"]
        df_sp.columns = df_sp.columns.astype(str).str.strip()
        st.session_state.sp_df = df_sp
        st.success(f"✅ 报表读取成功，共 {len(df_sp)} 行投放数据")
        st.markdown("### 原始数据预览（前10行）")
        st.dataframe(df_sp.head(10), use_container_width=True)
    except Exception as e:
        st.error(f"文件读取失败：{str(e)}")
        st.stop()

# ===================== 数据筛选工具函数 =====================
def filter_data_by_rules(df, filter_rules):
    mask = np.ones(len(df), dtype=bool)
    for rule in filter_rules:
        col_name = rule["col"].strip()
        op = rule["op"]
        val = rule["val"].strip()
        if not col_name or not val or col_name not in df.columns:
            continue
        # 数值转换
        try:
            num_val = float(val)
            df_col = pd.to_numeric(df[col_name], errors="coerce")
        except:
            num_val = None
            df_col = df[col_name].astype(str).str.strip()

        # 运算符逻辑处理
        if op == "等于":
            mask = mask & (df_col == num_val if num_val is not None else df_col == val)
        elif op == "大于" and num_val is not None:
            mask = mask & (df_col > num_val)
        elif op == "小于" and num_val is not None:
            mask = mask & (df_col < num_val)
        elif op == "大于等于" and num_val is not None:
            mask = mask & (df_col >= num_val)
        elif op == "小于等于" and num_val is not None:
            mask = mask & (df_col <= num_val)
        elif op == "包含":
            mask = mask & df_col.str.contains(val, na=False)
    return df[mask].copy()

# ===================== 报表生成主逻辑 =====================
if st.session_state.sp_df is not None:
    st.markdown("---")
    st.subheader("当前待生成Sheet清单预览")
    sheet_info_text = ""
    for idx, cfg in enumerate(st.session_state.sheet_config_list):
        sheet_info_text += f"{idx+1}. Sheet名称：{cfg['sheet_name']} | 输出字段数量：{len(cfg['output_cols'])}\n"
    st.text_area("配置汇总", value=sheet_info_text, height=150, disabled=True)

    run_btn = st.button("🚀 一键生成全部自定义Excel报表", type="primary", use_container_width=True)
    if run_btn:
        with st.spinner("正在生成Excel报表，请勿操作页面，完成后自动出现下载按钮"):
            df_origin = st.session_state.sp_df.copy()
            output_buf = io.BytesIO()
            try:
                with pd.ExcelWriter(output_buf, engine="openpyxl", mode="w") as writer:
                    for sheet_cfg in st.session_state.sheet_config_list:
                        sheet_name = sheet_cfg["sheet_name"]
                        filter_rules = sheet_cfg["filter_rules"]
                        output_cols = sheet_cfg["output_cols"]
                        match_asin = sheet_cfg["match_campaign_asin_sku"]
                        sort_by = sheet_cfg["sort_by"]
                        sort_asc = sheet_cfg["sort_ascending"]
                        is_targeting = sheet_cfg["is_targeting_sheet"]
                        only_match_bid = sheet_cfg["only_match_bidding"]

                        if not sheet_name or len(output_cols) == 0:
                            st.warning(f"跳过无效配置Sheet：{sheet_name}（未填写名称或输出字段）")
                            continue

                        ws = writer.book.create_sheet(sheet_name)
                        # 写入表头
                        for col_idx, col_name in enumerate(output_cols, start=1):
                            cell = ws.cell(row=1, column=col_idx, value=col_name)
                            cell.font = HEADER_FONT
                            cell.fill = HEADER_FILL
                            cell.alignment = CENTER_ALIGN
                            cell.border = THIN_BORDER

                        filter_df = filter_data_by_rules(df_origin, filter_rules)
                        final_rows = []

                        # 按Campaign匹配ASIN/SKU逻辑
                        if match_asin:
                            for _, line in filter_df.iterrows():
                                if "Campaign ID" not in df_origin.columns:
                                    final_rows.append(line)
                                    continue
                                cid = line["Campaign ID"]
                                match_group = df_origin[df_origin["Campaign ID"] == cid]
                                asin_sku_pairs = match_group[["ASIN (Informational only)", "SKU"]].dropna().drop_duplicates()
                                if len(asin_sku_pairs) == 0:
                                    continue
                                # 匹配字段处理
                                campaign_bid = match_group["Bidding Strategy"].dropna().iloc[0] if not match_group["Bidding Strategy"].dropna().empty else ""
                                campaign_state = match_group["Campaign State (Informational only)"].dropna().iloc[0] if not match_group["Campaign State (Informational only)"].dropna().empty else ""
                                ad_group_state = match_group["State"].dropna().iloc[0] if not match_group["State"].dropna().empty else ""
                                ad_group_name = match_group["Ad Group Name (Informational only)"].dropna().iloc[0] if not match_group["Ad Group Name (Informational only)"].dropna().empty else ""
                                # 精准广告Sheet补充Keyword Text匹配
                                keyword_text = match_group["Keyword Text"].dropna().iloc[0] if "Keyword Text" in match_group.columns and not match_group["Keyword Text"].dropna().empty else ""

                                for _, pair in asin_sku_pairs.iterrows():
                                    new_line = line.copy()
                                    new_line["ASIN (Informational only)"] = pair["ASIN (Informational only)"]
                                    new_line["SKU"] = pair["SKU"]
                                    new_line["Bidding Strategy"] = campaign_bid
                                    new_line["Campaign State (Informational only)"] = campaign_state
                                    new_line["State"] = ad_group_state
                                    new_line["Ad Group Name (Informational only)"] = ad_group_name
                                    new_line["Keyword Text"] = keyword_text
                                    # 仅匹配Bidding模式下清空Placement、Percentage
                                    if only_match_bid:
                                        new_line["Placement"] = ""
                                        new_line["Percentage"] = ""
                                    final_rows.append(new_line)
                        else:
                            final_rows = filter_df.to_dict("records")

                        df_result = pd.DataFrame(final_rows)
                        if sort_by in df_result.columns:
                            df_result = df_result.sort_values(by=sort_by, ascending=sort_asc)
                        st.info(f"【{sheet_name}】有效数据行数：{len(df_result)}")

                        # 定位超链接列索引
                        targeting_col_idx = None
                        if is_targeting and "Product Targeting Expression" in output_cols:
                            targeting_col_idx = output_cols.index("Product Targeting Expression") + 1

                        # 填充数据行
                        for row_idx, row_data in enumerate(df_result.to_dict("records"), start=2):
                            for col_idx, col_name in enumerate(output_cols, start=1):
                                val = row_data.get(col_name, "")
                                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                                cell.border = THIN_BORDER
                                cell.alignment = CENTER_ALIGN

                                # ASIN超链接处理（仅定位广告Sheet执行）
                                if is_targeting and col_idx == targeting_col_idx and val != "":
                                    asin_val = extract_asin_from_targeting(val)
                                    if asin_val:
                                        excel_formula = f'=HYPERLINK("https://www.amazon.com/dp/{asin_val}","{asin_val}")'
                                        cell.value = excel_formula
                                        cell.font = Font(color='0563C1', underline='single')
                                        cell.alignment = LEFT_ALIGN

                                # 数字格式处理
                                if val != "" and isinstance(val, (int, float)):
                                    if col_name in ["ACOS", "Click-through Rate", "Conversion Rate"]:
                                        cell.number_format = PCT_FMT
                                    elif col_name in ["Spend", "Sales", "CPC", "ROAS", "Percentage", "Bid"]:
                                        cell.number_format = NUMBER_FMT

                        auto_fit_columns(ws)
                        ws.freeze_panes = "A2"
                        del df_result, final_rows, filter_df
                    del df_origin

                output_buf.seek(0)
                st.success("✅ 全部自定义Sheet报表生成完成！")
                st.download_button(
                    label="📥 下载完整Excel文件",
                    data=output_buf.getvalue(),
                    file_name=f"SP广告自定义报表_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            except Exception as err:
                st.error(f"生成失败，错误详情：{str(err)}")